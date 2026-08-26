from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import TestCase
from openpyxl import load_workbook

from comparator.models import InventoryItem, MetroOffer, Product, StockMovement
from comparator.services.weekly_reports import build_weekly_report, build_weekly_report_xlsx


class WeeklyReportTests(TestCase):
    def setUp(self):
        self.staff = get_user_model().objects.create_user(
            username="report-admin",
            password="A-test-password-2026!",
            is_staff=True,
        )
        self.client.force_login(self.staff)
        self.product = Product.objects.create(name="Apă raport 2L", base_unit="L")
        self.inventory = InventoryItem.objects.create(
            product=self.product,
            minimum_quantity=Decimal("5"),
            target_quantity=Decimal("12"),
            retail_price_gross=Decimal("5"),
            retail_unit_size=Decimal("2"),
        )
        StockMovement.objects.create(
            inventory_item=self.inventory,
            quantity_delta=Decimal("2"),
            reason=StockMovement.Reason.OPENING,
        )
        MetroOffer.objects.create(
            product=self.product,
            units_per_package=1,
            unit_size=2,
            price_gross=Decimal("4"),
            valid_from=date(2026, 8, 20),
            source="METRO Târgoviște Punct",
        )

    def test_report_recommends_low_stock_and_builds_excel(self):
        report = build_weekly_report(date(2026, 8, 26))
        workbook = load_workbook(BytesIO(build_weekly_report_xlsx(report)))

        self.assertEqual(report["start_date"], date(2026, 8, 20))
        self.assertEqual(len(report["low_stock"]), 1)
        self.assertEqual(report["low_stock"][0]["needed"], Decimal("10"))
        self.assertEqual(report["low_stock"][0]["best"]["source"], "METRO Târgoviște Punct")
        self.assertEqual(
            workbook.sheetnames,
            ["Rezumat", "Reaprovizionare", "Marje de verificat", "Comparații săptămână", "Anomalii METRO"],
        )

    def test_private_web_report_and_export(self):
        page = self.client.get("/app/rapoarte/saptamanal/?end=2026-08-26")
        export = self.client.get("/app/rapoarte/saptamanal/export/?end=2026-08-26")

        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Apă raport 2L")
        self.assertEqual(export.status_code, 200)
        self.assertTrue(export.content.startswith(b"PK"))

    def test_command_writes_private_weekly_file(self):
        with TemporaryDirectory() as directory:
            call_command("generate_weekly_report", end="2026-08-26", output=Path(directory), verbosity=0)
            output = Path(directory) / "raport-saptamanal-2026-08-26.xlsx"
            self.assertTrue(output.exists())
            self.assertEqual(output.stat().st_mode & 0o777, 0o600)
