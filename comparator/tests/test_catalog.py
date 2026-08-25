import io
from datetime import date
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import TestCase
from openpyxl import load_workbook

from comparator.catalog import infer_category
from comparator.models import MetroOffer, MetroOfferTier, MetroScrapeJob, MetroScrapedProduct, Product


class CategoryTests(TestCase):
    def test_infers_representative_categories(self):
        self.assertEqual(infer_category("Ciocolata cu lapte 100 g"), "Dulciuri")
        self.assertEqual(infer_category("Pasta de dinti mentol 75 ml"), "Igienă personală")
        self.assertEqual(infer_category("Detergent vase lamaie 1 L"), "Curățenie")
        self.assertEqual(infer_category("Salam uscat 500 g"), "Mezeluri")

    def test_categorize_command_uses_keywords_then_staging_fallback(self):
        job = MetroScrapeJob.objects.create(start_url="https://produse.metro.ro/shop")
        detergent = Product.objects.create(name="Detergent vase Mere 1 L", base_unit="L")
        unknown = Product.objects.create(name="Brand fără indicii 100 g", base_unit="KG")
        common = {
            "job": job,
            "product_url": "https://produse.metro.ro/shop/pv/test",
            "store_name": "METRO PUNCT TARGOVISTE",
            "units_per_package": 1,
            "unit_size": 1,
            "base_unit": "KG",
            "price_gross": Decimal("5.00"),
            "imported": True,
        }
        MetroScrapedProduct.objects.create(
            **common,
            external_id="D1",
            name=detergent.name,
            category="Fructe și legume",
            matched_product=detergent,
        )
        MetroScrapedProduct.objects.create(
            **common,
            external_id="U1",
            name=unknown.name,
            category="Snacks",
            matched_product=unknown,
        )
        call_command("categorize_products", "--overwrite")
        detergent.refresh_from_db()
        unknown.refresh_from_db()
        self.assertEqual(detergent.category, "Curățenie")
        self.assertEqual(unknown.category, "Snacks")


class CatalogExportTests(TestCase):
    def setUp(self):
        self.staff = get_user_model().objects.create_user(
            username="export-admin", password="A-test-password-2026!", is_staff=True
        )
        self.client.force_login(self.staff)
        self.product = Product.objects.create(
            name="Chipsuri test 100 g", category="Snacks", base_unit="KG"
        )
        offer = MetroOffer.objects.create(
            product=self.product,
            units_per_package=1,
            unit_size=Decimal("0.1"),
            price_gross=Decimal("5.50"),
            valid_from=date(2026, 7, 15),
            source="Selenium METRO PUNCT TARGOVISTE",
        )
        MetroOfferTier.objects.create(
            offer=offer,
            min_packages=4,
            price_gross=Decimal("4.80"),
        )

    def test_catalog_can_be_filtered_by_category(self):
        Product.objects.create(name="Lapte ascuns 1 L", category="Lactate", base_unit="L")
        response = self.client.get("/app/catalog/?category=Snacks")
        self.assertContains(response, "Chipsuri test")
        self.assertNotContains(response, "Lapte ascuns")

    def test_excel_export_contains_catalog_and_offer_history(self):
        response = self.client.get("/app/catalog/export/xlsx/?category=Snacks")
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Catalog curent", "Toate ofertele"])
        self.assertEqual(workbook["Catalog curent"]["A2"].value, "Chipsuri test 100 g")
        self.assertEqual(workbook["Catalog curent"]["D2"].value, "Snacks")
        self.assertIn("4+ pachete", workbook["Catalog curent"]["J2"].value)

    def test_csv_export_is_excel_compatible(self):
        response = self.client.get("/app/catalog/export/csv/?category=Snacks")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b"\xef\xbb\xbf"))
        self.assertIn("Chipsuri test 100 g".encode(), response.content)
        self.assertIn(b";", response.content)
        self.assertIn(b"4+ pachete", response.content)
