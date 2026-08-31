import io
from datetime import date
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import load_workbook

from comparator.models import (
    ActivityLog,
    Invoice,
    InvoiceLine,
    InventoryItem,
    MetroOffer,
    MetroOfferTier,
    PriceAlert,
    Product,
    ShoppingList,
    ShoppingListItem,
    StockMovement,
    Supplier,
    SupplierOffer,
)


class CompleteDataExportTests(TestCase):
    def setUp(self):
        users = get_user_model()
        self.owner = users.objects.create_superuser(
            "export-owner", password="A-test-password-2026!"
        )
        self.operator = users.objects.create_user(
            "export-operator", password="A-test-password-2026!", is_staff=True
        )
        self.supplier = Supplier.objects.create(name="=Furnizor formulă", tax_id="RO123")
        self.product = Product.objects.create(
            name="=WEBSERVICE(\"https://invalid.example\")",
            brand="Test",
            ean="4006381333931",
            category="Snacks",
            base_unit="BUC",
        )
        self.offer = MetroOffer.objects.create(
            product=self.product,
            price_gross=Decimal("8.50"),
            valid_from=date(2026, 8, 20),
            source="METRO Târgoviște",
        )
        MetroOfferTier.objects.create(
            offer=self.offer,
            min_packages=4,
            price_gross=Decimal("7.50"),
        )
        self.invoice = Invoice.objects.create(
            supplier=self.supplier,
            number="F-100",
            issued_at=date(2026, 8, 21),
            document_total_gross=Decimal("10.00"),
        )
        self.line = InvoiceLine.objects.create(
            invoice=self.invoice,
            original_name="@Produs factură",
            quantity=1,
            unit_price_gross=Decimal("10.00"),
            matched_product=self.product,
            match_score=100,
            match_method=InvoiceLine.MatchMethod.MANUAL,
            needs_review=False,
        )
        SupplierOffer.objects.create(
            supplier=self.supplier,
            product=self.product,
            invoice_line=self.line,
            price_per_base_unit=Decimal("10.00"),
            base_unit="BUC",
            valid_from=date(2026, 8, 21),
        )
        inventory = InventoryItem.objects.create(
            product=self.product,
            minimum_quantity=2,
            target_quantity=5,
            retail_price_gross=Decimal("12.00"),
        )
        StockMovement.objects.create(
            inventory_item=inventory,
            quantity_delta=Decimal("3"),
            reason=StockMovement.Reason.OPENING,
            source_key="export-test-opening",
            created_by=self.owner,
        )
        shopping_list = ShoppingList.objects.create(name="Lista test", budget_gross=100)
        ShoppingListItem.objects.create(
            shopping_list=shopping_list,
            product=self.product,
            quantity=2,
        )
        PriceAlert.objects.create(product=self.product, target_price=Decimal("8.00"))

    def test_complete_export_contains_operational_sheets_and_safe_cells(self):
        self.client.force_login(self.owner)

        response = self.client.get(
            "/app/exporturi/descarca/?start_date=2026-08-01&end_date=2026-08-31"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment", response["Content-Disposition"])
        workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=False)
        self.assertEqual(
            workbook.sheetnames,
            [
                "Rezumat",
                "Produse",
                "Prețuri METRO",
                "Praguri METRO",
                "Furnizori",
                "Prețuri furnizori",
                "Documente",
                "Linii documente",
                "Inventar",
                "Mișcări stoc",
                "Liste cumpărături",
                "Alerte",
            ],
        )
        self.assertEqual(workbook["Produse"]["B2"].value, "'=WEBSERVICE(\"https://invalid.example\")")
        self.assertEqual(workbook["Furnizori"]["B2"].value, "'=Furnizor formulă")
        self.assertEqual(workbook["Linii documente"]["E2"].value, "'@Produs factură")
        self.assertEqual(workbook["Praguri METRO"]["C2"].value, 4)
        self.assertEqual(workbook["Inventar"]["D2"].value, 3)

    def test_date_filter_excludes_old_transactional_rows(self):
        MetroOffer.objects.create(
            product=self.product,
            price_gross=Decimal("9.00"),
            valid_from=date(2025, 1, 1),
            source="METRO vechi",
        )
        self.client.force_login(self.owner)

        response = self.client.get(
            "/app/exporturi/descarca/?start_date=2026-08-01&end_date=2026-08-31"
        )
        workbook = load_workbook(io.BytesIO(response.content), read_only=True)

        metro_rows = list(workbook["Prețuri METRO"].iter_rows(values_only=True))
        self.assertEqual(len(metro_rows), 2)
        self.assertEqual(metro_rows[1][9], "METRO Târgoviște")

    def test_download_is_admin_only_and_audited(self):
        self.client.force_login(self.operator)
        self.assertEqual(self.client.get("/app/exporturi/").status_code, 403)
        self.assertEqual(self.client.get("/app/exporturi/descarca/").status_code, 403)
        self.client.force_login(self.owner)

        response = self.client.get("/app/exporturi/descarca/")

        self.assertEqual(response.status_code, 200)
        event = ActivityLog.objects.get(
            user=self.owner,
            method="GET",
            view_name="comparator:data_export_download",
        )
        self.assertEqual(event.outcome, ActivityLog.Outcome.SUCCESS)

    def test_invalid_interval_returns_validation_error_without_export(self):
        self.client.force_login(self.owner)

        response = self.client.get(
            "/app/exporturi/descarca/?start_date=2026-09-01&end_date=2026-08-01"
        )

        self.assertEqual(response.status_code, 400)
        self.assertContains(
            response,
            "Data de început nu poate fi după data de sfârșit.",
            status_code=400,
        )
