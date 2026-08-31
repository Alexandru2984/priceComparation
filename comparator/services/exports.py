import csv
import io
from datetime import datetime
from decimal import Decimal

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from comparator.models import (
    InventoryItem,
    Invoice,
    InvoiceLine,
    MetroOffer,
    MetroOfferTier,
    PriceAlert,
    Product,
    ShoppingListItem,
    StockMovement,
    Supplier,
    SupplierOffer,
)

from .inventory import inventory_with_balance

CATALOG_HEADERS = [
    "Produs", "Marcă", "EAN", "Categorie", "Unitate bază", "Preț pachet", "Bucăți/pachet",
    "Cantitate/bucată", "Preț/unitate bază", "Praguri volum", "Cel mai mic preț/unitate",
    "Sursă", "Valabil de la", "Activ",
]


SPREADSHEET_FORMULA_PREFIXES = (
    "=", "+", "-", "@", "\t", "\r", "\n", "＝", "＋", "－", "＠",
)


def _safe_cell(value):
    """Keep user-controlled text from becoming a spreadsheet formula."""
    if isinstance(value, str) and value.startswith(SPREADSHEET_FORMULA_PREFIXES):
        return f"'{value}"
    if isinstance(value, datetime) and timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    return value


def _format_workbook(workbook):
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="195C44")
        for column in sheet.columns:
            letter = column[0].column_letter
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 55)
            sheet.column_dimensions[letter].width = max(width, 12)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, Decimal):
                    cell.value = float(cell.value)
                    cell.number_format = "0.0000"
                elif isinstance(cell.value, datetime):
                    cell.number_format = "dd.mm.yyyy hh:mm"


def _append_sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append([_safe_cell(value) for value in row])
    return sheet


def _filter_date_range(queryset, field, start_date, end_date):
    if start_date:
        queryset = queryset.filter(**{f"{field}__gte": start_date})
    if end_date:
        queryset = queryset.filter(**{f"{field}__lte": end_date})
    return queryset


def _catalog_row(product):
    offer = product.current_metro_offer()
    tiers = list(offer.volume_tiers.all()) if offer else []
    tier_summary = " | ".join(
        f"{tier.min_packages}+ pachete: {tier.price_gross:.2f} lei"
        for tier in tiers
    )
    lowest_price = min(
        [offer.price_per_base_unit, *(tier.price_per_base_unit for tier in tiers)]
    ) if offer else None
    return [
        product.name,
        product.brand,
        product.ean,
        product.category or "Altele",
        product.base_unit,
        offer.price_gross if offer else None,
        offer.units_per_package if offer else None,
        offer.unit_size if offer else None,
        offer.price_per_base_unit if offer else None,
        tier_summary,
        lowest_price,
        offer.source if offer else "",
        offer.valid_from if offer else None,
        "Da" if product.active else "Nu",
    ]


def build_catalog_csv(products):
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(CATALOG_HEADERS)
    for product in products:
        writer.writerow([_safe_cell(value) for value in _catalog_row(product)])
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def build_catalog_xlsx(products, offers):
    workbook = Workbook()
    catalog = workbook.active
    catalog.title = "Catalog curent"
    catalog.append(CATALOG_HEADERS)
    for product in products:
        catalog.append([_safe_cell(value) for value in _catalog_row(product)])

    history = workbook.create_sheet("Toate ofertele")
    history_headers = [
        "Produs", "Categorie", "Unitate bază", "Preț pachet", "Bucăți/pachet", "Cantitate/bucată",
        "Preț/unitate bază", "Praguri volum", "Cel mai mic preț/unitate", "Sursă", "Valabil de la", "Activă",
    ]
    history.append(history_headers)
    for offer in offers:
        tiers = list(offer.volume_tiers.all())
        history.append(
            [
                _safe_cell(offer.product.name),
                offer.product.category or "Altele",
                offer.product.base_unit,
                offer.price_gross,
                offer.units_per_package,
                offer.unit_size,
                offer.price_per_base_unit,
                " | ".join(
                    f"{tier.min_packages}+ pachete: {tier.price_gross:.2f} lei"
                    for tier in tiers
                ),
                min([offer.price_per_base_unit, *(tier.price_per_base_unit for tier in tiers)]),
                offer.source,
                offer.valid_from,
                "Da" if offer.active else "Nu",
            ]
        )

    _format_workbook(workbook)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_complete_data_xlsx(*, start_date=None, end_date=None, include_inactive=False):
    """Build a portable operational export without secrets, OCR text or uploaded files."""
    products = Product.objects.all().order_by("name", "brand")
    offers = MetroOffer.objects.select_related("product").order_by("product__name", "-valid_from", "id")
    tiers = MetroOfferTier.objects.select_related("offer", "offer__product").order_by(
        "offer__product__name", "offer_id", "min_packages"
    )
    inventory = InventoryItem.objects.select_related("product").order_by("product__name")
    shopping_items = ShoppingListItem.objects.select_related("shopping_list", "product").order_by(
        "shopping_list_id", "product__name"
    )
    alerts = PriceAlert.objects.select_related("product").order_by("product__name")
    if not include_inactive:
        products = products.filter(active=True)
        offers = offers.filter(active=True)
        tiers = tiers.filter(offer__active=True)
        inventory = inventory.filter(active=True)
        shopping_items = shopping_items.filter(shopping_list__archived=False)
        alerts = alerts.filter(active=True)

    offers = _filter_date_range(offers, "valid_from", start_date, end_date)
    tiers = _filter_date_range(tiers, "offer__valid_from", start_date, end_date)
    supplier_offers = _filter_date_range(
        SupplierOffer.objects.select_related("supplier", "product", "invoice_line__invoice").order_by(
            "product__name", "-valid_from"
        ),
        "valid_from",
        start_date,
        end_date,
    )
    invoices = _filter_date_range(
        Invoice.objects.select_related("supplier").prefetch_related("lines").order_by("issued_at", "id"),
        "issued_at",
        start_date,
        end_date,
    )
    invoice_lines = _filter_date_range(
        InvoiceLine.objects.select_related("invoice", "invoice__supplier", "matched_product").order_by(
            "invoice__issued_at", "invoice_id", "id"
        ),
        "invoice__issued_at",
        start_date,
        end_date,
    )
    movements = StockMovement.objects.select_related(
        "inventory_item__product", "invoice_line__invoice", "sale_line", "created_by"
    ).order_by("created_at", "id")
    if start_date:
        movements = movements.filter(created_at__date__gte=start_date)
        shopping_items = shopping_items.filter(shopping_list__created_at__date__gte=start_date)
        alerts = alerts.filter(created_at__date__gte=start_date)
    if end_date:
        movements = movements.filter(created_at__date__lte=end_date)
        shopping_items = shopping_items.filter(shopping_list__created_at__date__lte=end_date)
        alerts = alerts.filter(created_at__date__lte=end_date)

    workbook = Workbook()
    workbook.remove(workbook.active)
    exported_at = timezone.localtime()
    interval = f"{start_date or 'început'} – {end_date or 'prezent'}"
    summary_rows = [
        ("Generat la", exported_at),
        ("Interval", interval),
        ("Include inactive/arhivate", "Da" if include_inactive else "Nu"),
        ("Produse", products.count()),
        ("Oferte METRO", offers.count()),
        ("Praguri METRO", tiers.count()),
        ("Furnizori", Supplier.objects.count()),
        ("Prețuri furnizori", supplier_offers.count()),
        ("Documente", invoices.count()),
        ("Linii documente", invoice_lines.count()),
        ("Produse în inventar", inventory.count()),
        ("Mișcări de stoc", movements.count()),
        ("Articole în liste", shopping_items.count()),
        ("Alerte", alerts.count()),
        ("Date excluse", "parole, MFA, abonamente push, text OCR, fișiere și căi locale"),
    ]
    _append_sheet(workbook, "Rezumat", ["Câmp", "Valoare"], summary_rows)
    _append_sheet(
        workbook,
        "Produse",
        ["ID", "Produs", "Marcă", "EAN", "Categorie", "Unitate bază", "Activ", "Creat", "Actualizat"],
        (
            (p.pk, p.name, p.brand, p.ean, p.category or "Altele", p.base_unit, "Da" if p.active else "Nu", p.created_at, p.updated_at)
            for p in products
        ),
    )
    _append_sheet(
        workbook,
        "Prețuri METRO",
        ["ID", "Produs", "EAN", "Bucăți/pachet", "Cantitate/bucată", "Unitate", "Preț pachet", "Preț/unitate", "Valabil de la", "Sursă", "Activ"],
        (
            (o.pk, o.product.name, o.product.ean, o.units_per_package, o.unit_size, o.product.base_unit, o.price_gross, o.price_per_base_unit, o.valid_from, o.source, "Da" if o.active else "Nu")
            for o in offers
        ),
    )
    _append_sheet(
        workbook,
        "Praguri METRO",
        ["ID ofertă", "Produs", "De la pachete", "Preț/pachet", "Preț/unitate", "Etichetă", "Valabil de la", "Sursă"],
        (
            (t.offer_id, t.offer.product.name, t.min_packages, t.price_gross, t.price_per_base_unit, t.label, t.offer.valid_from, t.offer.source)
            for t in tiers
        ),
    )
    _append_sheet(
        workbook,
        "Furnizori",
        ["ID", "Denumire", "CUI", "METRO", "Comandă minimă", "Transport", "Transport gratuit de la", "Observații"],
        (
            (s.pk, s.name, s.tax_id, "Da" if s.is_metro else "Nu", s.minimum_order_gross, s.transport_gross, s.free_transport_from, s.notes)
            for s in Supplier.objects.order_by("name")
        ),
    )
    _append_sheet(
        workbook,
        "Prețuri furnizori",
        ["Furnizor", "Produs", "EAN", "Preț/unitate", "Unitate", "Valabil de la", "Document"],
        (
            (o.supplier.name, o.product.name, o.product.ean, o.price_per_base_unit, o.base_unit, o.valid_from, o.invoice_line.invoice.number)
            for o in supplier_offers
        ),
    )
    _append_sheet(
        workbook,
        "Documente",
        ["ID", "Furnizor", "Tip", "Număr", "Data", "Status", "Recepționat în stoc", "Transport", "Reducere document", "Total declarat", "Total calculat", "Reconciliat", "Linii", "Observații"],
        (
            (i.pk, i.supplier.name, i.get_document_type_display(), i.number, i.issued_at, i.get_status_display(), "Da" if i.receive_into_stock else "Nu", i.transport_gross, i.document_discount_gross, i.document_total_gross, i.calculated_document_total_gross, "Da" if i.is_reconciled else "Nu", len(i.lines.all()), i.notes)
            for i in invoices
        ),
    )
    _append_sheet(
        workbook,
        "Linii documente",
        ["ID document", "Data", "Furnizor", "Număr", "Denumire originală", "EAN", "Cantitate", "Bucăți/pachet", "Cantitate/bucată", "Unitate", "Preț brut", "TVA %", "Total linie", "Reducere", "SGR", "Produs asociat", "Scor", "Metodă", "De verificat", "Preț/unitate efectiv"],
        (
            (line.invoice_id, line.invoice.issued_at, line.invoice.supplier.name, line.invoice.number, line.original_name, line.ean, line.quantity, line.units_per_package, line.unit_size, line.base_unit, line.unit_price_gross, line.vat_rate, line.calculated_line_total, line.discount_gross, line.deposit_gross, line.matched_product.name if line.matched_product else "", line.match_score, line.get_match_method_display(), "Da" if line.needs_review else "Nu", line.price_per_base_unit)
            for line in invoice_lines
        ),
    )
    inventory_rows = inventory_with_balance(inventory)
    _append_sheet(
        workbook,
        "Inventar",
        ["Produs", "EAN", "Unitate", "Stoc curent", "Minim", "Țintă", "Sub minim", "Valabilitate zile", "Preț raft", "Cantitate/unitate vândută", "TVA vânzare", "TVA achiziție", "Marjă țintă %", "Pierderi %", "Activ"],
        (
            (item.product.name, item.product.ean, item.product.base_unit, item.current_quantity, item.minimum_quantity, item.target_quantity, "Da" if item.is_low else "Nu", item.shelf_life_days, item.retail_price_gross, item.retail_unit_size, item.retail_vat_rate, item.purchase_vat_rate, item.target_margin_percent, item.expected_waste_percent, "Da" if item.active else "Nu")
            for item in inventory_rows
        ),
    )
    _append_sheet(
        workbook,
        "Mișcări stoc",
        ["Data", "Produs", "Modificare", "Unitate", "Motiv", "Document", "Referință vânzare", "Cheie sursă", "Notă", "Utilizator"],
        (
            (m.created_at, m.inventory_item.product.name, m.quantity_delta, m.inventory_item.product.base_unit, m.get_reason_display(), m.invoice_line.invoice.number if m.invoice_line_id else "", m.sale_line.external_reference if m.sale_line_id else "", m.source_key, m.note, m.created_by.username if m.created_by_id else "")
            for m in movements
        ),
    )
    _append_sheet(
        workbook,
        "Liste cumpărături",
        ["ID listă", "Listă", "Creată", "Buget", "Arhivată", "Produs", "EAN", "Cantitate", "Unitate", "Prioritate", "Cumpărat"],
        (
            (item.shopping_list_id, item.shopping_list.name, item.shopping_list.created_at, item.shopping_list.budget_gross, "Da" if item.shopping_list.archived else "Nu", item.product.name, item.product.ean, item.quantity, item.product.base_unit, item.get_priority_display(), "Da" if item.purchased else "Nu")
            for item in shopping_items
        ),
    )
    _append_sheet(
        workbook,
        "Alerte",
        ["Produs", "EAN", "Prag", "Unitate", "Notă", "Activă", "Creată", "Ultima notificare", "Ultimul preț notificat"],
        (
            (a.product.name, a.product.ean, a.target_price, a.product.base_unit, a.note, "Da" if a.active else "Nu", a.created_at, a.last_notified_at, a.last_notified_price)
            for a in alerts
        ),
    )
    _format_workbook(workbook)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
