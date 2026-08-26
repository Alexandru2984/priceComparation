import io
from datetime import timedelta
from decimal import Decimal

from django.db.models import Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from comparator.models import Invoice, InvoiceLine, InventoryItem, MetroPriceAnomaly, StockMovement

from .exports import _safe_cell
from .insights import current_source_options, profitability_summary
from .inventory import inventory_with_balance


def build_weekly_report(end_date=None):
    end_date = end_date or timezone.localdate()
    start_date = end_date - timedelta(days=6)
    documents = list(
        Invoice.objects.select_related("supplier")
        .prefetch_related("lines")
        .filter(issued_at__range=(start_date, end_date))
    )
    confirmed_lines = InvoiceLine.objects.select_related(
        "invoice", "invoice__supplier", "matched_product"
    ).prefetch_related("matched_product__metro_offers__volume_tiers").filter(
        invoice__issued_at__range=(start_date, end_date),
        needs_review=False,
        matched_product__isnull=False,
    )
    comparison_rows = []
    for line in confirmed_lines:
        comparison = line.comparison()
        if comparison:
            comparison_rows.append((line, comparison))

    inventory = list(inventory_with_balance(InventoryItem.objects.filter(active=True)))
    low_stock = []
    for item in inventory:
        needed = item.replenishment_quantity
        if needed <= 0:
            continue
        options = current_source_options(item.product, needed)
        best = options[0] if options else None
        worst = options[-1] if options else None
        low_stock.append({
            "item": item,
            "current": item.current_quantity,
            "needed": needed,
            "best": best,
            "saving": worst["total"] - best["total"] if best and worst and len(options) > 1 else Decimal("0"),
        })
    low_stock.sort(key=lambda row: (row["current"] - row["item"].minimum_quantity, row["item"].product.name))

    margins = profitability_summary(inventory)
    margin_rows = [
        (item, analysis)
        for item, analysis in margins["rows"]
        if analysis["status"] in {"LOSS", "BELOW_TARGET", "INCOMPLETE"}
    ]
    status_order = {"LOSS": 0, "BELOW_TARGET": 1, "INCOMPLETE": 2}
    margin_rows.sort(key=lambda row: (status_order[row[1]["status"]], row[0].product.name))

    sales = StockMovement.objects.filter(
        reason=StockMovement.Reason.SALE,
        sale_line__sold_at__date__range=(start_date, end_date),
    )
    sales_quantity = -(sales.aggregate(total=Sum("quantity_delta"))["total"] or Decimal("0"))
    anomalies = list(
        MetroPriceAnomaly.objects.select_related("product")
        .filter(status=MetroPriceAnomaly.Status.OPEN)
        .order_by("-detected_at")[:100]
    )
    extra_cost = sum(
        (comparison["total_impact"] for _, comparison in comparison_rows if comparison["total_impact"] > 0),
        Decimal("0"),
    )
    savings = sum(
        (-comparison["total_impact"] for _, comparison in comparison_rows if comparison["total_impact"] < 0),
        Decimal("0"),
    )
    return {
        "start_date": start_date,
        "end_date": end_date,
        "documents": documents,
        "document_count": len(documents),
        "purchase_total": sum((document.calculated_document_total_gross for document in documents), Decimal("0")),
        "comparison_rows": comparison_rows,
        "extra_cost": extra_cost,
        "savings": savings,
        "low_stock": low_stock,
        "margin_rows": margin_rows,
        "margin_summary": margins,
        "anomalies": anomalies,
        "sales_count": sales.count(),
        "sales_quantity": sales_quantity,
    }


def build_weekly_report_xlsx(report):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Rezumat"
    summary.append(["Raport", "Valoare"])
    summary_rows = [
        ("Perioadă", f"{report['start_date']:%d.%m.%Y}–{report['end_date']:%d.%m.%Y}"),
        ("Documente achiziție", report["document_count"]),
        ("Achiziții totale (lei)", report["purchase_total"]),
        ("Economii față de METRO (lei)", report["savings"]),
        ("Cost suplimentar față de METRO (lei)", report["extra_cost"]),
        ("Produse sub minim", len(report["low_stock"])),
        ("Marje sub țintă/pierdere/incomplete", len(report["margin_rows"])),
        ("Anomalii METRO deschise", len(report["anomalies"])),
        ("Mișcări de vânzare", report["sales_count"]),
        ("Cantitate de bază vândută", report["sales_quantity"]),
    ]
    for row in summary_rows:
        summary.append([_safe_cell(value) for value in row])

    stock = workbook.create_sheet("Reaprovizionare")
    stock.append(["Produs", "Stoc", "Minim", "Țintă", "De cumpărat", "Sursă recomandată", "Cost estimat", "Economie posibilă"])
    for row in report["low_stock"]:
        stock.append([
            _safe_cell(row["item"].product.name), row["current"], row["item"].minimum_quantity,
            row["item"].target_quantity, row["needed"], row["best"]["source"] if row["best"] else "",
            row["best"]["total"] if row["best"] else None, row["saving"],
        ])

    margins = workbook.create_sheet("Marje de verificat")
    margins.append(["Produs", "Status", "Cost net efectiv", "Preț vânzare", "Marjă %", "Preț recomandat"])
    for item, analysis in report["margin_rows"]:
        margins.append([
            _safe_cell(item.product.name), analysis["status"], analysis["effective_cost_net"],
            analysis["retail_gross"], analysis["margin_percent"], analysis["recommended_retail_gross"],
        ])

    comparisons = workbook.create_sheet("Comparații săptămână")
    comparisons.append(["Document", "Furnizor", "Produs", "Cost furnizor", "Cost METRO", "Impact total", "Status"])
    for line, analysis in report["comparison_rows"]:
        comparisons.append([
            line.invoice.number, _safe_cell(line.invoice.supplier.name), _safe_cell(line.original_name),
            line.price_per_base_unit, analysis["metro_price"], analysis["total_impact"], analysis["status"],
        ])

    anomalies = workbook.create_sheet("Anomalii METRO")
    anomalies.append(["Produs extras", "Preț anterior", "Preț nou", "Abatere %", "Detectată"])
    for anomaly in report["anomalies"]:
        anomalies.append([
            _safe_cell(anomaly.product.name), anomaly.old_price_per_base, anomaly.new_price_per_base,
            anomaly.change_percent, anomaly.detected_at.replace(tzinfo=None),
        ])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="195C44")
        for column in sheet.columns:
            letter = column[0].column_letter
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 55)
            sheet.column_dimensions[letter].width = max(width, 12)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, Decimal):
                    cell.value = float(cell.value)
                    cell.number_format = "0.00"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
