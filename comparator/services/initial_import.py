import hashlib
import io
import re
import unicodedata
import zipfile
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from comparator.catalog import CATEGORY_CHOICES, infer_category
from comparator.models import BaseUnit, InitialDataImport, InventoryItem, Product, StockMovement, Supplier

from .barcodes import assign_ean, is_valid_gtin, normalize_barcode


MAX_ROWS = 5000
MAX_XLSX_UNCOMPRESSED = 50 * 1024 * 1024
CATEGORIES = {value for value, _ in CATEGORY_CHOICES}
CATEGORY_BY_NORMALIZED_NAME = {}


def _normal(value):
    value = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return " ".join(re.sub(r"[^a-z0-9]+", " ", value.lower()).split())


CATEGORY_BY_NORMALIZED_NAME.update({_normal(category): category for category in CATEGORIES})


def _text(value, limit=240):
    if value is None:
        return ""
    result = str(value).strip()
    if result.endswith(".0") and result[:-2].isdigit():
        result = result[:-2]
    return result[:limit]


def _decimal(value, default=None):
    if value in (None, ""):
        return default
    cleaned = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _boolean(value):
    return _normal(value) in {"1", "da", "yes", "true", "x"}


def _unit(value):
    normalized = _normal(value)
    mapping = {
        "buc": BaseUnit.PIECE,
        "bucata": BaseUnit.PIECE,
        "bucati": BaseUnit.PIECE,
        "kg": BaseUnit.KILOGRAM,
        "kilogram": BaseUnit.KILOGRAM,
        "l": BaseUnit.LITER,
        "litru": BaseUnit.LITER,
        "litri": BaseUnit.LITER,
    }
    return mapping.get(normalized)


def _records(sheet):
    iterator = sheet.iter_rows(values_only=True)
    headers = next(iterator, None)
    if not headers:
        return []
    headers = [_normal(value) for value in headers]
    rows = []
    for row_number, values in enumerate(iterator, start=2):
        record = {header: value for header, value in zip(headers, values) if header}
        if any(value not in (None, "") for value in record.values()):
            rows.append((row_number, record))
    return rows


def _value(record, *aliases):
    for alias in aliases:
        if alias in record:
            return record[alias]
    return None


def _supplier_rows(sheet):
    result = []
    for number, record in _records(sheet):
        name = _text(_value(record, "denumire", "nume", "furnizor"), 180)
        errors = [] if name else ["Denumire furnizor lipsă"]
        values = {
            "name": name,
            "tax_id": _text(_value(record, "cui", "cod fiscal"), 30),
            "is_metro": _boolean(_value(record, "metro", "este metro")),
            "minimum_order_gross": _decimal(_value(record, "comanda minima"), Decimal("0")),
            "transport_gross": _decimal(_value(record, "transport"), Decimal("0")),
            "free_transport_from": _decimal(_value(record, "transport gratuit de la")),
            "notes": _text(_value(record, "observatii", "note"), 1000),
        }
        for field in ("minimum_order_gross", "transport_gross"):
            if values[field] is None or values[field] < 0:
                errors.append(f"Valoare invalidă pentru {field}")
        if values["free_transport_from"] is not None and values["free_transport_from"] < 0:
            errors.append("Prag transport gratuit invalid")
        result.append({"kind": "SUPPLIER", "sheet": sheet.title, "row": number, "data": values, "errors": errors})
    return result


def _product_rows(sheet):
    result = []
    seen_eans = set()
    for number, record in _records(sheet):
        name = _text(_value(record, "denumire", "nume", "produs"), 220)
        brand = _text(_value(record, "marca", "brand"), 100)
        ean = normalize_barcode(_text(_value(record, "ean", "gtin", "cod de bare"), 30))
        unit = _unit(_value(record, "unitate", "unitate baza", "um"))
        supplied_category = _text(_value(record, "categorie"), 80)
        category = CATEGORY_BY_NORMALIZED_NAME.get(_normal(supplied_category)) if supplied_category else None
        category = category or (infer_category(f"{name} {brand}") if not supplied_category else supplied_category)
        errors = []
        if not name:
            errors.append("Denumire produs lipsă")
        if unit is None:
            errors.append("Unitatea trebuie să fie BUC, KG sau L")
        if ean and not is_valid_gtin(ean):
            errors.append("EAN/GTIN invalid")
        if ean and (ean in seen_eans or Product.objects.filter(ean=ean).exclude(name__iexact=name).exists()):
            errors.append("EAN duplicat")
        if ean:
            seen_eans.add(ean)
        if category not in CATEGORIES:
            errors.append("Categorie necunoscută")
        values = {
            "name": name,
            "brand": brand,
            "ean": ean,
            "base_unit": unit or "",
            "category": category,
            "active": not _normal(_value(record, "activ")) in {"nu", "0", "false"},
        }
        result.append({"kind": "PRODUCT", "sheet": sheet.title, "row": number, "data": values, "errors": errors})
    return result


def _stock_rows(sheet, product_rows):
    workbook_eans = {row["data"]["ean"] for row in product_rows if row["data"]["ean"] and not row["errors"]}
    workbook_names = {_normal(row["data"]["name"]) for row in product_rows if not row["errors"]}
    result = []
    for number, record in _records(sheet):
        name = _text(_value(record, "denumire", "nume", "produs"), 220)
        ean = normalize_barcode(_text(_value(record, "ean", "gtin", "cod de bare"), 30))
        errors = []
        if not ean and not name:
            errors.append("Completează EAN sau denumirea produsului")
        if ean and ean not in workbook_eans and not Product.objects.filter(ean=ean).exists():
            errors.append("EAN-ul nu există în catalog sau în foaia Produse")
        if not ean and name:
            matches = Product.objects.filter(name__iexact=name).count()
            if _normal(name) not in workbook_names and matches == 0:
                errors.append("Produsul nu există în catalog sau în foaia Produse")
            elif matches > 1:
                errors.append("Denumire ambiguă; completează EAN")
        values = {
            "name": name,
            "ean": ean,
            "opening_quantity": _decimal(_value(record, "stoc initial"), Decimal("0")),
            "minimum_quantity": _decimal(_value(record, "stoc minim"), Decimal("0")),
            "target_quantity": _decimal(_value(record, "stoc tinta"), Decimal("0")),
            "retail_price_gross": _decimal(_value(record, "pret vanzare")),
            "retail_unit_size": _decimal(_value(record, "cantitate unitate vanduta"), Decimal("1")),
            "retail_vat_rate": _decimal(_value(record, "tva vanzare"), Decimal("0")),
            "purchase_vat_rate": _decimal(_value(record, "tva achizitie"), Decimal("0")),
            "target_margin_percent": _decimal(_value(record, "marja tinta"), Decimal("20")),
            "expected_waste_percent": _decimal(_value(record, "pierderi estimate"), Decimal("0")),
        }
        for field in ("opening_quantity", "minimum_quantity", "target_quantity"):
            if values[field] is None or values[field] < 0:
                errors.append(f"Valoare invalidă pentru {field}")
        if values["minimum_quantity"] is not None and values["target_quantity"] is not None:
            if values["target_quantity"] < values["minimum_quantity"]:
                errors.append("Stocul țintă trebuie să fie cel puțin stocul minim")
        if values["retail_price_gross"] is not None and values["retail_price_gross"] < 0:
            errors.append("Preț de vânzare invalid")
        if values["retail_unit_size"] is None or values["retail_unit_size"] <= 0:
            errors.append("Cantitatea unității vândute trebuie să fie pozitivă")
        for field in ("retail_vat_rate", "purchase_vat_rate"):
            if values[field] is None or not 0 <= values[field] <= 100:
                errors.append(f"Valoare invalidă pentru {field}")
        for field in ("target_margin_percent", "expected_waste_percent"):
            if values[field] is None or not 0 <= values[field] < 100:
                errors.append(f"Valoare invalidă pentru {field}")
        result.append({"kind": "STOCK", "sheet": sheet.title, "row": number, "data": values, "errors": errors})
    return result


def parse_initial_workbook(upload):
    data = upload.read()
    upload.seek(0)
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            if sum(item.file_size for item in archive.infolist()) > MAX_XLSX_UNCOMPRESSED:
                raise ValueError("Registrul XLSX se extinde la peste 50 MB și a fost refuzat.")
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except (OSError, ValueError, zipfile.BadZipFile) as exc:
        raise ValueError("Registrul XLSX nu este valid.") from exc
    sheets = {_normal(name): workbook[name] for name in workbook.sheetnames}
    missing = [name for name in ("furnizori", "produse", "stoc") if name not in sheets]
    if missing:
        raise ValueError("Lipsesc foile obligatorii: " + ", ".join(missing) + ".")
    suppliers = _supplier_rows(sheets["furnizori"])
    products = _product_rows(sheets["produse"])
    stock = _stock_rows(sheets["stoc"], products)
    rows = [*suppliers, *products, *stock]
    workbook.close()
    if not rows:
        raise ValueError("Registrul nu conține niciun rând de importat.")
    if len(rows) > MAX_ROWS:
        raise ValueError(f"Registrul poate avea maximum {MAX_ROWS} rânduri.")
    for row in rows:
        row["data"] = {key: str(value) if isinstance(value, Decimal) else value for key, value in row["data"].items()}
    return hashlib.sha256(data).hexdigest(), rows


def build_initial_workbook_template():
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instrucțiuni"
    instructions.append(["PriceMatch · import inițial"])
    instructions.append(["Completează foile Furnizori, Produse și Stoc. Pentru asocierea sigură a stocului folosește EAN."])
    sheets = {
        "Furnizori": ["Denumire", "CUI", "METRO", "Comanda minima", "Transport", "Transport gratuit de la", "Observatii"],
        "Produse": ["Denumire", "Marca", "EAN", "Unitate", "Categorie", "Activ"],
        "Stoc": ["EAN", "Denumire", "Stoc initial", "Stoc minim", "Stoc tinta", "Pret vanzare", "Cantitate unitate vanduta", "TVA vanzare", "TVA achizitie", "Marja tinta", "Pierderi estimate"],
    }
    for title, headers in sheets.items():
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{chr(64 + min(len(headers), 26))}1"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="195C44")
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = 22
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _find_product(data):
    if data.get("ean"):
        return Product.objects.filter(ean=data["ean"]).first()
    products = Product.objects.filter(name__iexact=data.get("name", ""))
    if "brand" in data:
        products = products.filter(brand__iexact=data.get("brand", ""))
    if data.get("base_unit"):
        products = products.filter(base_unit=data["base_unit"])
    return products.order_by("id").first()


@transaction.atomic
def apply_initial_import(initial_import):
    initial_import = InitialDataImport.objects.select_for_update().get(pk=initial_import.pk)
    if initial_import.status == InitialDataImport.Status.APPLIED:
        return {"already_applied": True}
    stats = {"suppliers_created": 0, "suppliers_updated": 0, "products_created": 0, "products_updated": 0, "stock_policies": 0, "opening_movements": 0, "skipped": sum(bool(row["errors"]) for row in initial_import.rows)}

    for row in initial_import.rows:
        if row["kind"] != "SUPPLIER" or row["errors"]:
            continue
        data = row["data"]
        supplier = Supplier.objects.filter(name__iexact=data["name"]).first()
        created = supplier is None
        supplier = supplier or Supplier(name=data["name"])
        for field in ("tax_id", "is_metro", "minimum_order_gross", "transport_gross", "free_transport_from", "notes"):
            setattr(supplier, field, data[field])
        supplier.full_clean()
        supplier.save()
        stats["suppliers_created" if created else "suppliers_updated"] += 1

    for row in initial_import.rows:
        if row["kind"] != "PRODUCT" or row["errors"]:
            continue
        data = row["data"]
        product = _find_product(data)
        if product is None:
            product = Product(name=data["name"], brand=data["brand"], base_unit=data["base_unit"])
            created = True
        else:
            created = False
        product.name = data["name"]
        product.brand = data["brand"]
        product.base_unit = data["base_unit"]
        product.category = data["category"]
        product.active = data["active"]
        product.full_clean(exclude=["ean"])
        product.save()
        if data["ean"]:
            assign_ean(product, data["ean"])
        stats["products_created" if created else "products_updated"] += 1

    for row in initial_import.rows:
        if row["kind"] != "STOCK" or row["errors"]:
            continue
        data = row["data"]
        product = _find_product(data)
        if product is None:
            stats["skipped"] += 1
            continue
        inventory, _ = InventoryItem.objects.update_or_create(
            product=product,
            defaults={
                "minimum_quantity": Decimal(data["minimum_quantity"]),
                "target_quantity": Decimal(data["target_quantity"]),
                "retail_price_gross": Decimal(data["retail_price_gross"])
                if data["retail_price_gross"] not in (None, "None", "")
                else None,
                "retail_unit_size": Decimal(data["retail_unit_size"]),
                "retail_vat_rate": Decimal(data["retail_vat_rate"]),
                "purchase_vat_rate": Decimal(data["purchase_vat_rate"]),
                "target_margin_percent": Decimal(data["target_margin_percent"]),
                "expected_waste_percent": Decimal(data["expected_waste_percent"]),
                "active": True,
            },
        )
        stats["stock_policies"] += 1
        opening = Decimal(data["opening_quantity"])
        if opening > 0:
            _, created = StockMovement.objects.update_or_create(
                source_key=f"INITIAL:{initial_import.pk}:{row['row']}",
                defaults={
                    "inventory_item": inventory,
                    "quantity_delta": opening,
                    "reason": StockMovement.Reason.OPENING,
                    "note": f"Import inițial din {initial_import.original_filename}"[:240],
                    "created_by": initial_import.created_by,
                },
            )
            stats["opening_movements"] += int(created)

    initial_import.status = InitialDataImport.Status.APPLIED
    initial_import.applied_at = timezone.now()
    initial_import.save(update_fields=["status", "applied_at"])
    return stats
